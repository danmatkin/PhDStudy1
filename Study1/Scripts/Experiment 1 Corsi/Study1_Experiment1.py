from tkinter import *
import openpyxl
from tkmacosx import *
import time
import sys
import random
import shutil
import glob
import os.path
from PIL import Image, ImageTk
from PIL.Image import Resampling
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart  # Multipurpose internet mail extensions
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from pathlib import Path
from cryptography.fernet import Fernet
from itertools import tee
import smtplib

DEV_MODE = FAST_COMPONENT_DEMO = FULL_COMPONENT_DEMO = False
EXPERIMENT = CONTROL = False
q_path = 'Q:/Research/PSP/PhD_DanMatkin/study_1_ex_1_top_up.xlsx.encrypted'

if not os.path.isfile(q_path):
    modes_dict = {
        's': (True, False, False),
        'c': (False, False, True),
        'f': (False, True, False)
    }
    comic_dict = {
        'a': 'A',
        'b': 'B',
        'c': 'C'
    }
    mode = input('\nSecure drive unreachable.\n\nChoose run mode:\n   s - Small Window, Fast, Dev Mode\n   c - Full Component Demo\n   f - Fast Component Demo\n >>> ')
    if mode in modes_dict:
        DEV_MODE, FAST_COMPONENT_DEMO, FULL_COMPONENT_DEMO = modes_dict[mode]

    run_type = input('\nChoose run type:\n   e - Experiment Run\n   c - Control Run\n >>> ')
    if run_type == 'e':
        comics_set = input('\nChoose Comics Set:\n   a - A\n   b - B\n   c - C\n >>> ')

root = Tk()
root.configure(background='White')
SCREEN_WIDTH = root.winfo_screenwidth()
SCREEN_HEIGHT = root.winfo_screenheight()

if DEV_MODE:
    WINDOW_WIDTH = int(SCREEN_WIDTH / 2.5)
    WINDOW_HEIGHT = int(SCREEN_HEIGHT / 2.5)
    root.geometry(f'{WINDOW_WIDTH}x{WINDOW_HEIGHT}+{int(SCREEN_HEIGHT / 1)}+0')  # Set window size
else:
    WINDOW_WIDTH = SCREEN_WIDTH
    WINDOW_HEIGHT = SCREEN_HEIGHT

def save_data(data_workbook, responses):
    WORKSHEET.append(responses)
    if os.path.isfile(q_path):
        data_workbook.save('Q:/Research/PSP/PhD_DanMatkin/study_1_ex_1_top_up.xlsx')
        close_datafile(KEY)
    else:
        data_workbook.save(f'{ID_CODE}_Experiment1_DemoData.xlsx')

def close_datafile(key):
    """Encrypt the datafile and remove the unencrypted datafile"""
    with open('Q:/Research/PSP/PhD_DanMatkin/study_1_ex_1_top_up.xlsx', 'rb') as data_file:
        data = data_file.read()
    fernet = Fernet(key)
    encrypted = fernet.encrypt(data)
    data_file.write(encrypted)
    os.remove('Q:/Research/PSP/PhD_DanMatkin/study_1_ex_1_top_up.xlsx')
    email_protocol()

def create_canvas_comics(images_list):
    """This function places all comics onto a tkinter canvas for presentation and appends them to a list."""
    comics_list = []
    for comic in images_list:
        comic_canvas = Canvas(root, width=WINDOW_WIDTH, height=WINDOW_HEIGHT, bg='white', )
        comic_canvas.create_image(int(WINDOW_WIDTH / 2 + 3), int(WINDOW_HEIGHT / 2 + 3), image=comic)
        comics_list.append(comic_canvas)
    return comics_list

def create_comics(window_height):
    """Load and resize the images into a new folder,
    sort them back into their original order,
    remove the folder,
    return the list of images."""

    initial_image_list = []
    image_height = int(window_height / 3)  # Determine the height of the images as a third of the screen size.

    new_comics_path = resource_path('ResizedComics')

    os.makedirs(new_comics_path)
    for filename in glob.glob(f'{resource_path("Comics")}/*.jpg'): # For every file in this folder:
        image = Image.open(filename) # Open the file
        width_height = image.size # get the width/height of the image
        width_old = width_height[0] # Specify the old width
        height_old = width_height[1] # Specify the old Height
        width_new = (image_height * width_old) / height_old # Generate numbers for reizing the images.
        image_resized = image.resize((int(width_new), int(image_height)), Resampling.LANCZOS) # Resize the images using the
        image_resized.save('{}{}{}'.format(new_comics_path, '/', os.path.split(filename)[1])) # Save the

    for file in sorted(glob.glob(f'{resource_path("ResizedComics")}/*.jpg')):  # Sort the images
        images = ImageTk.PhotoImage(Image.open(file))
        initial_image_list.append(images)

    shutil.rmtree(new_comics_path)

    # image_list = [initial_image_list[i:i + 3] for i in range(0, len(initial_image_list), 3)]
    return initial_image_list

def create_comic_dictionary(images_list):
    """This function organises the experiment comics into a dictionary,
    this is to get easy access to the comic identifier."""
    comic_values = ('Arrow3','Arrow4', 'Arrow5', 'Boomerang3', 'Boomerang4', 'Boomerang5', 'Boxing3', 'Boxing4',
                    'Boxing5', 'Bus3', 'Bus4', 'Bus5', 'Dog3', 'Dog4', 'Dog5', 'Sailor3', 'Sailor4', 'Sailor5')
    comics_dictionary = dict(zip(comic_values, images_list))
    return comics_dictionary

def create_practice_comic(practice_comics):
    """This function creates the practice comic"""
    practice = random.choice(practice_comics) # pick a random comic
    # make a canvas to put the comic on.
    base = Canvas(root, width=WINDOW_WIDTH, height=WINDOW_HEIGHT, bg='white')
    # Make a canvas to put the comic on.
    practice_comic = Canvas(base, width=int(practice.width() + 3), height=int(practice.height() + 3), bg='White', highlightthickness=0, bd=0)
    # Create the image of the comic, with the width and height of image + 3 and set the image to practice.
    practice_comic.create_image(int(practice.width() /2 + 3), int(practice.height()/2 + 3), image=practice)
    # place the practice comic onto the
    practice_comic.place(relx=.5, rely=.5, anchor=CENTER)
    return base

def email_protocol():
    """Writes master_responses to the data file,
    sends the follow up email to the participant."""



    email_sender = 'comicsmemoryexperiment@gmail.com'
    email_password = 'ngbfsmjqdrzkxycq'
    email_reciever = RESPONSES[6]

    subject = 'Comics Experiment Follow Up'

    body = f"""Hi there! 

Thank you for participating in the comics study, attached to this email are the information, consent and debrief materials you saw during the experiment.

Your participant ID code is {RESPONSES[3]}.

If for any reason at all you want your data to be removed from the study, you should get in touch with me at daniel.matkin@shu.ac.uk WITHIN 24 HOURS of recieving this email.

Following this 24 hour period the data you provided will be anonymised.


The VLFI is intended to measure your fluency with the comic book medium, average scores on this survey are around 12, you scored {RESPONSES[9]}.

If you have any questions either about the experiment, then please feel free to get in touch.

Thanks again for participating, it's very much appreciated!

All the best,
Dan"""

    message = MIMEMultipart()
    message['from'] = email_sender
    message['to'] = email_reciever
    message['subject'] = subject
    message.attach(MIMEText(body))
    message.attach(MIMEApplication(Path(resource_path('ParticipantInformationMaterials.pdf')).read_bytes()))

    with smtplib.SMTP(host='smtp.gmail.com', port=587) as smtp:
        smtp.ehlo()
        smtp.starttls()  # Transport layer security
        smtp.login(email_sender, email_password)
        smtp.send_message(message)

def quit_protocol():
    path = 'Q:/Research/PSP/PhD_DanMatkin/study2_experiment1.xlsx'
    if os.path.isfile(path):
        os.remove(path)
    return sys.exit()

def esc_protocol():
    """Quit function. This file is called at the end of the protocol.
        Necessary if participant presses escape key."""
    path = 'Q:/Research/PSP/PhD_DanMatkin/study2_experiment1.xlsx'
    if os.path.isfile(path):
        os.remove(path)
    else:
        os.remove(resource_path(f'{ID_CODE}_Experiment1_DemoData.xlsx'))
    return sys.exit()

def font(screen_height, font_size):
    """This function handles font sizes so they remain constant."""
    for n in range(1, 100):
        x = int(screen_height / n)
        if x - font_size == 0:
            return n
        elif x - font_size in range(-3, 0):
            return n
        elif x - font_size in range(-4, 0):
            return n

def get_comic_set(data):
    """This function chooses the set of comics to run"""
    global EXPERIMENT, CONTROL
    if EXPERIMENT:
        if data[0] == 'A':
            this_run_comics = 'B'
        elif data[0] == 'B':
            this_run_comics = 'C'
        else:
            this_run_comics = 'A'
        RESPONSES.append(this_run_comics)
        return this_run_comics
    elif CONTROL:
        RESPONSES.append('Control')

def get_comic_names(ver, comics):
    """This function uses the comic dictionary to assign the correct identifier to the comic."""
    questions = ['What facial hair if any, did the archer have?',
                 'What did the man do while waiting for the boomerang to come back?',
                 'How many ropes did the boxing ring have?',
                 'What number was one the side of the bus?',
                 'What pattern was on the dog walker\'s jumper?',
                 'How many portholes did the sailor\'s ship have?']
    if ver == 'A':
        comic_names = ['Arrow3','Boomerang4','Boxing5','Bus3','Dog4','Sailor5']
    elif ver == 'B':
        comic_names = ['Arrow4','Boomerang5','Boxing3','Bus4','Dog5','Sailor3']
    else:
        comic_names = ['Arrow5','Boomerang3','Boxing4','Bus5','Dog3','Sailor4']
    zipped = list(zip(comics, comic_names, questions))
    random.shuffle(zipped)
    comics, names, questions = zip(*zipped)
    comics, names, questions = list(comics), list(names), list(questions)
    return comics, names, questions

def get_comics(images_list, ver):
    """This function takes gets the comic block for this experiment and makes them ready for the experiment."""
    comic_list = []
    if ver == 'A':
        image_index = 0
        for _ in range(0, 6):
            comic_list.append(images_list[image_index])
            if len(comic_list) == 3:
                image_index = 9
            else:
                image_index += 4
    elif ver == 'B':
        image_index = 1
        for _ in range(0, 6):
            comic_list.append(images_list[image_index])
            if len(comic_list) == 2:
                image_index = 6
            elif len(comic_list) == 5:
                image_index = 15
            else:
                image_index += 4
    elif ver == 'C':
        image_index = 2
        for _ in range(0, 6):
            comic_list.append(images_list[image_index])
            if len(comic_list) == 1:
                image_index = 3
            elif len(comic_list) == 4:
                image_index = 12
            else:
                image_index += 4
    return comic_list

def get_practice_comics(images_list):
    """This comic gets the practice comic ready for presentation."""
    practice_comics = []
    practice_comics.extend(images_list[3:6])
    del images_list[3:6]
    return practice_comics

def open_key():
    """This function opens the key for the data file."""
    with open(resource_path('key_study_1_ex_1_top_up.key'), 'rb') as key_file:  # Provides the key to open the datafile
        key = key_file.read()
    return key

def open_datafile(key):
    """Use the key to decrypt the datafile"""

    with open('Q:/Research/PSP/PhD_DanMatkin/study_1_ex_1_top_up.xlsx.encrypted', 'rb') as data_file:
        data = data_file.read()

    fernet = Fernet(key)
    decrypted = fernet.decrypt(data)

    with open('Q:/Research/PSP/PhD_DanMatkin/study_1_ex_1_top_up.xlsx', 'wb') as data_file:
        data_file.write(decrypted)

def check_id(data_worksheet):
    """Generates the unique ID Code for the participant.
    ensures the code is unique."""
    def generate_code():
        alphanumeric = ''.join(random.choice('0123456789abcdef') for _ in range(6))
        return alphanumeric

    id_code = generate_code()
    identifiers = [cell.value for cell in data_worksheet['D']]
    while id_code in identifiers:
        id_code = generate_code()

    return id_code

def generate_id():
    alphanumeric = ''.join(random.choice('0123456789abcdef') for _ in range(6))
    return alphanumeric

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = getattr(sys, '_MEIPASS', os.getcwd())
    except Exception:
        base_path = os.path.abspath("")
    return os.path.join(base_path, relative_path)

root.bind('<Escape>', lambda esc: esc_protocol()) # Binds escape key to quit function.

RESPONSES = []

if os.path.isfile(q_path): # Protocol initiated for if a file is set up.
    KEY = open_key()
    open_datafile(KEY)

    DATA_WORKBOOK = load_workbook(q_path)
    WORKSHEET = DATA_WORKBOOK.worksheets[0]

    if WORKSHEET.max_row % 2 == 0:  # if the number of rows in the sheet is even:
        EXPERIMENT = True
    else:
        CONTROL = True

    last_data_run = []
    if EXPERIMENT:
        last_row_idx = WORKSHEET.max_row
        last_data_run = [cell.value for cell in WORKSHEET[last_row_idx - 1]]
    else:
        for _ in WORKSHEET.iter_rows():
            last_data_run = [_.value for _ in _]

    ID_CODE = check_id(WORKSHEET)
    THIS_RUN_VERSION = get_comic_set(last_data_run)

    if EXPERIMENT:
        INITIAL_IMAGES_LIST = create_comics(WINDOW_HEIGHT)  # Find and resize all the images.
        PRACTICE_COMICS = get_practice_comics(INITIAL_IMAGES_LIST)  # Cut the practice comics out of the images list.
        COMICS_BLOCK = get_comics(INITIAL_IMAGES_LIST, THIS_RUN_VERSION)  # Get the list of comics that we want.
        COMICS, COMIC_NAMES, QUESTIONS = get_comic_names(THIS_RUN_VERSION, COMICS_BLOCK)
        COMICS = create_canvas_comics(COMICS)
        PRACTICE_COMIC = create_practice_comic(PRACTICE_COMICS)

else: # Protocol for Demo Mode
    if run_type == 'e':
        EXPERIMENT = True
        RESPONSES.append(comic_dict[comics_set])
    else:
        CONTROL = True
        RESPONSES.append('Control')
    ID_CODE = generate_id()
    DATA_WORKBOOK = openpyxl.Workbook()
    DATA_WORKBOOK.save(f'{ID_CODE}_Experiment1_DemoData.xlsx')
    WORKSHEET = DATA_WORKBOOK.worksheets[0]

    if EXPERIMENT:
        INITIAL_IMAGES_LIST = create_comics(WINDOW_HEIGHT)  # Find and resize all the images.
        PRACTICE_COMICS = get_practice_comics(INITIAL_IMAGES_LIST)  # Cut the practice comics out of the images list.
        COMICS_BLOCK = get_comics(INITIAL_IMAGES_LIST, comic_dict[comics_set]) # Get the list of comics that we want.
        COMICS, COMIC_NAMES, QUESTIONS = get_comic_names(comic_dict[comics_set], COMICS_BLOCK)
        COMICS = create_canvas_comics(COMICS)
        PRACTICE_COMIC = create_practice_comic(PRACTICE_COMICS)

if DEV_MODE or FAST_COMPONENT_DEMO:
    two_seconds = 200
    seven_seconds = 700
    pre_presentation_time = 200
    one_second = 100
    two_hundred_milliseconds = 20
    ten_seconds = 1000
else:
    two_seconds = 2000
    seven_seconds = 7000
    pre_presentation_time = 2000
    one_second = 1000
    two_hundred_milliseconds = 200
    ten_seconds = 10000  # Define the timings used in the procedure.

class InformationSheet:

    def __init__(self, window):
        # -------------------------------- Information Sheet Initialisations --------------------------------
        logo_height = int(WINDOW_HEIGHT / 6)
        logo_path = resource_path('cebsap.png')
        logo = Image.open(logo_path)
        width_height = logo.size
        width_old = width_height[0]
        height_old = width_height[1]
        width_new = (logo_height * width_old) / height_old
        logo_resized = logo.resize((int(width_new), int(logo_height)), Resampling.LANCZOS)
        logo_resized.save('cebsapNew.png')
        # The above code initiliases the SHU Logo that goes onto the top of the information sheet.
        # The logo_height variable sets how tall the logo should be and then the rest is set based on that.
        if not DEV_MODE:
            window.attributes('-fullscreen', True)  # Set the window to fullscreen.

        self.window = window  # Window variable for easier manipulation later.
        self.logo = ImageTk.PhotoImage(Image.open('cebsapNew.png'))  # The logo in a variable for later use.
        os.remove('cebsapNew.png')

        self.page1 = Canvas(window, bg='white', width=int(WINDOW_WIDTH / 2), height=WINDOW_HEIGHT)
        self.page2 = Canvas(window, bg='white', width=int(WINDOW_WIDTH / 2),
                            height=WINDOW_HEIGHT)  # Canvases for each side of the information sheet.
        self.consent_page = Canvas(window, bg='white', width=int(WINDOW_WIDTH / 2),
                                   height=WINDOW_HEIGHT)  # Consent Page canvas.

        # -------------------------------- Information Sheet Text Variables --------------------------------
        self.intro_text = Label(self.page1, bg='white',
                                text="""Thank you for considering participating in this study!\n\nPlease read the following information carefully.""",
                                font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10)), 'bold'),
                                justify='center')
        self.purpose_heading = Label(self.page1, bg='white', text="""1. What's the aim of the study?""",
                                     font=('Calibri', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), 'bold'),
                                     justify='left',
                                     anchor='w')
        self.purpose_text = Label(self.page1,
                                  bg='white', text="""
                The aim of this study is to investigate the use of significant visuospatial working memory capacity during 
        comics reading.""",
                                  font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))),
                                  justify='left')
        self.participation_heading = Label(self.page1, bg='white', text="""2. Do I have to take part?""",
                                           font=('Calibri', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), 'bold'),
                                           justify='left',
                                           anchor='w')
        self.participation_text = Label(self.page1, bg='white', text="""
                Taking part in this study is entirely voluntary. If you begin the experiment and decide you want to stop at any 
        point, you are completely free to do so with no questions asked.

                The same applies to if you finish the experiment and decide that you want to withdraw your data. All you need to
        do is email the lead researcher at daniel.matkin@shu.ac.uk with your unique alphanumeric code that you will be 
        given on the next page, indicating that you want your data withdrawn and the data will be destroyed with no questions 
        asked. Once the experiment is over, you'll receive all these details again, so there's no need to worry about remembering 
        them right now.""", font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))), justify='left')
        self.tasks_heading = Label(self.page1, bg='white', text='3. What do I have to do?',
                                   font=('Calibri', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), 'bold'),
                                   justify='left', anchor='w')
        self.tasks_text = Label(self.page1, bg='white', text="""
                There are 3 parts to this experiment. Firstly you’ll be asked to complete a baseline working memory test. 
        When that section is complete, you’ll be asked to a survey about your previous experience of and familiarity with comics. 
        Finally, you’ll be asked to complete a number of trials of the same working memory task you completed but you’ll 
        do this while viewing comics or doing other secondary tasks.""", font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))), justify='left')
        self.confidentiality_heading = Label(self.page1, bg='white', text='4. Will my data be confidential?',
                                             font=('Calibri', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), 'bold'),
                                             justify='left',
                                             anchor='w')
        self.confidentiality_text = Label(self.page1, bg='white', text="""
                The data that you provide here will be kept completely confidential. The only identifying features on the data 
        will be the randomly generated alphanumeric code that will you be provided with on the next page. Only you will know your
        code, meaning that it will be impossible for anyone else to ever be trace your data back to you. The only reason that the 
        data will be accessed during the data collection period will be to remove data if requested.

                 On the next page you will also be asked to provide an email address to which a copy of this information will be 
        sent, as well as any follow up information that you may request. These email addresses will be stored in an encrypted 
        file separate to the file containing the data.""", font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))),
                                          justify='left')
        self.confidentiality2text = Label(self.page2, bg='white', text="""
                        Once you have completed the experiment, the data will be written automatically into a symmetrically encrypted, 
                password protected spreadsheet that only the lead researcher and Director of Studies has the access key to.

                        After the data collection period has finished, the alphanumeric IDs and separate email addresses file will be 
                permanently erased, anonymising the data in line with the UK GDPR best practice principles. Please note, that should you 
                wish to withdraw your data from the study, you should contact the researcher about removing your data before the 20th of December, 2022.""",
                                          font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))), justify='left')
        self.ethics_heading = Label(self.page2, bg='white', text='5. Has this study been ethically appraised?',
                                    font=('Calibri', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), 'bold'),
                                    justify='left', anchor='w')
        self.ethics_text = Label(self.page2, bg='white', text="""
                        The Sheffield Hallam University Ethics board has given approval for this study to go ahead and
                determined that there are no significant risks associated with participating in this research.""",
                                 font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))), justify='left')
        self.gdpr_heading = Label(self.page2, bg='white', text='6. General Data Protection Regulation ',
                                  font=('Calibri', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), 'bold'),
                                  justify='left', anchor='w')
        self.gdpr_text = Label(self.page2, bg='white', text="""
                        The GDPR governs the way in which organisations use your personal data. Personal data is information relating to 
                an identifiable living individual. Transparency is a key element of the GDPR and before consenting to this research you 
                should understand:

                -	How and why your data is being used for research
                -	What your rights are under GDPR 
                -	How to contact us if you have questions or concerns about the use of your personal data. 

                The University undertakes research as part of its function for the community under the legal status. Data protection allows 
                us to use personal data for research with appropriate safeguards in place under the legal basis of public tasks that are 
                in the public interest. 

                A full statement of you rights can be found at:
                https://www.shu.ac.uk/about-this-website/privacy-policy/privacy-notices/privacy-notice-for-research

                However, all University research is reviewed to ensure that participants are treated appropriately and their rights 
                respected. This study was approved by the Research Ethics Working Group of the Department of Psychology, Sociology and 
                Politics.""", font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))), justify='left')
        self.contact_heading = Label(self.page2, bg='white', text='7. Contact Details',
                                     font=('Calibri', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), 'bold'),
                                     justify='left', anchor='w')
        self.contact_text = Label(self.page2, bg='white', text="""
                        For questions on how the University uses your data or if you would like to complain about data usage then you should 
                contact the Data Protection Officer at DPO@shu.ac.uk.

                If you have any other concerns you should contact in the first instance you should contact any of my supervisory team:

                                    Dr Paul Aleixo (Director of Studies)         Dr Jane Morgan
                                    dspa5@exchange.shu.ac.uk                    ssljlm@exchange.shu.ac.uk

                                                                    Dr Diarmuid Verrier
                                                                    dsdv@exchange.shu.ac.uk
                """, font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))), justify='left')
        self.questions_heading = Label(self.page2, bg='white',
                                       text='If you have any questions at all, please take the opportunity to ask them now.',
                                       font=('Calibri', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), 'bold'),
                                       justify='center')
        # -------------------------------- Information Sheet Text Variables --------------------------------

        self.forward_button = Button(self.page2, text='>>', width=int((WINDOW_WIDTH / 100) * 5),
                                     height=int(WINDOW_WIDTH / 100 * 1.3), bd=1,
                                     command=self.forward)  # Button to move the information sheet forward.

        # -------------------------------- Consent Radio Button Variables --------------------------------
        self.read = IntVar()
        self.questions = IntVar()
        self.withdraw = IntVar()
        self.agree = IntVar()
        self.participate = IntVar()
        self.consent_var = IntVar()
        # -------------------------------- Consent Radio Button Variables --------------------------------
        self.consent_frame = Frame(self.consent_page, bg='white', highlightthickness=2, highlightbackground='black')
        self.continue_label = Label(self.consent_frame, bg='white',
                                    text='Please answer the following questions by ticking the boxes if you are happy to proceed.',
                                    font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))), wraplength=500,
                                    justify='center')
        self.read_button = Checkbutton(self.consent_frame, bg='white',
                                       text="I have read the Information Sheet for this study and have had details of the study explained to me.",
                                       variable=self.read, font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))),
                                       wraplength=500,
                                       justify='center')
        self.questions_button = Checkbutton(self.consent_frame, bg='white',
                                            text="""My questions about the study have been answered to my satisfaction and I understand that I may ask further questions at any point.""",
                                            variable=self.questions,
                                            font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))),
                                            wraplength=500, justify='center')
        self.withdraw_button = Checkbutton(self.consent_frame, bg='white',
                                           text=""" I understand that I am free to withdraw from the study within the time limits outlined in the Information Sheet, without giving a reason for my withdrawal or to decline to answer any particular questions in the study without any consequences to my future treatment by the researcher.""",
                                           variable=self.withdraw,
                                           font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))),
                                           wraplength=500, justify='center')
        self.agree_button = Checkbutton(self.consent_frame, bg='white',
                                        text="""I agree to provide information to the researchers under the conditions of confidentiality set out in the Information Sheet.""",
                                        variable=self.agree,
                                        font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))), wraplength=500,
                                        justify='center')
        self.participate_button = Checkbutton(self.consent_frame, bg='white',
                                              text="""I wish to participate in the study under the conditions set out in the Information Sheet.""",
                                              variable=self.participate,
                                              font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))),
                                              wraplength=500, justify='center')
        self.consent_button = Checkbutton(self.consent_frame, bg='white',
                                          text="""I consent to the information collected for the purposes of this research study, once anonymised (so that I cannot be identified), to be used for any other research purposes.""",
                                          variable=self.consent_var,
                                          font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))),
                                          wraplength=500, justify='center')
        self.warning_label = Label(self.consent_page, bg='white',
                                   text='To continue with the experiment, check the boxes above to give your consent to participate.',
                                   fg='red', font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))))
        # -------------------------------- Consent Radio Button Variables --------------------------------

        self.back_button = Button(self.consent_page, text='<<', width=int((WINDOW_WIDTH / 100) * 5),
                                  height=int(WINDOW_WIDTH / 100 * 1.3), bd=1, command=self.back)
        self.start_button = Button(self.consent_page, text='Start', width=int((WINDOW_WIDTH / 100) * 5),
                                   height=int(WINDOW_WIDTH / 100 * 1.3), bd=1,
                                   command=self.consent_check)  # Button that moves the program on to the demographics page.

        # -------------------------------- Demographics Initialisations --------------------------------
        self.illegals = ['Create a Unique User ID',
                         'Enter Your Age',
                         '',
                         'Unique User ID Required',
                         'Age Required',
                         'Must be at least 4 characters',
                         'Enter your Email',
                         'Email Required',
                         'Too young to do experiment',
                         'Valid Email Required']  # Things the demographics window will not accept.
        self.entry_canvas = Canvas(self.window, bg='white', width=int(WINDOW_WIDTH / 100) * 27,
                                   height=int(WINDOW_HEIGHT / 100) * 55)
        self.demographics_frame = Frame(self.entry_canvas, bg='white', highlightthickness=1,
                                        highlightbackground='black',
                                        height=int((WINDOW_HEIGHT / 100) * 32), width=int((WINDOW_WIDTH / 100) * 23.22))
        self.demographics_label = Label(self.entry_canvas, bg='white', text='Please enter the following information',
                                        font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))))
        self.id_label = Label(self.entry_canvas, bg='white', text=f'Your Unique Identifier is: {ID_CODE}',
                              font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 10))))  # ID Code
        # -------------------------------- Demographics Entry Box Variables --------------------------------
        self.age = Entry(self.entry_canvas, justify=CENTER)
        self.email = Entry(self.entry_canvas, justify=CENTER)
        # -------------------------------- Demographics Gender Variables --------------------------------
        self.submit_button = Button(self.entry_canvas, text='Continue', width=int((WINDOW_WIDTH / 100) * 5),
                                    height=int(WINDOW_WIDTH / 100 * 1.3), command=self.submit, bd=1)
        self.gender_options = ['Male', 'Female', 'Trans-Man', 'Trans-Woman', 'Non-Binary', 'Prefer not to say']
        self.selection = StringVar()
        self.selection.set('Female')
        self.gender = OptionMenu(self.entry_canvas, self.selection, *self.gender_options)
        # -------------------------------- Demographics Variable Initialisation --------------------------------

    def information_page(self):
        """This function places all of the elemnents of the information sheet.
        At some point I would like to find a way to do this:
        a) more efficiently
        b) Dynamically - So the items place after each other based on their size"""

        self.page1.create_image(0, 0, image=self.logo, anchor=NW)
        self.intro_text.place(relx=.5, rely=.24, anchor=CENTER)
        self.purpose_heading.place(relx=.05, rely=.31, anchor=W)
        self.purpose_text.place(relx=.07, rely=.32, anchor=NW)
        self.participation_heading.place(relx=.05, rely=.41, anchor=W)
        self.participation_text.place(relx=.07, rely=.42, anchor=NW)
        self.tasks_heading.place(relx=.05, rely=.6, anchor=W)
        self.tasks_text.place(relx=.07, rely=.61, anchor=NW)
        self.confidentiality_heading.place(relx=.05, rely=.77, anchor=W)
        self.confidentiality_text.place(relx=.07, rely=.78, anchor=NW)
        self.confidentiality2text.place(relx=.07, rely=.05, anchor=NW)
        self.ethics_heading.place(relx=.05, rely=.21, anchor=W)
        self.ethics_text.place(relx=.07, rely=.22, anchor=NW)
        self.gdpr_heading.place(relx=.05, rely=.3, anchor=W)
        self.gdpr_text.place(relx=.07, rely=.31, anchor=NW)
        self.contact_heading.place(relx=.05, rely=.65, anchor=W)
        self.contact_text.place(relx=.07, rely=.66, anchor=NW)
        self.questions_heading.place(relx=.5, rely=.9, anchor=CENTER)
        self.forward_button.place(relx=0.8, rely=0.96, anchor=CENTER)

        self.page1.place(relx=0, rely=0)
        self.page1.create_line(int(WINDOW_WIDTH / 2), 0, int(WINDOW_WIDTH / 2), WINDOW_HEIGHT,
                               width=2)  # This creates a tkinter line.
        self.page2.place(relx=0.5, rely=0, anchor=NW)

    def show_consent_page(self):
        """This function places all of the consent page elements.
        Again, I want to make this more dynamic."""

        self.window.config(bg='#e6e1e1')
        self.consent_page.create_image(0, 0, image=self.logo, anchor=NW)
        self.consent_page.place(relx=0.5, rely=0.5, anchor=CENTER)
        self.consent_frame.place(relx=0.5, rely=0.5, anchor=CENTER)
        self.continue_label.grid(row=1, column=0, sticky=N)
        self.read_button.grid(row=2, column=0, sticky=N, pady=10)
        self.questions_button.grid(row=3, column=0, sticky=N, pady=10)
        self.withdraw_button.grid(row=4, column=0, sticky=N, pady=10)
        self.agree_button.grid(row=5, column=0, sticky=N, pady=10)
        self.participate_button.grid(row=6, column=0, sticky=N, pady=10)
        self.consent_button.grid(row=7, column=0, sticky=N, pady=10)
        self.back_button.place(relx=0.197, rely=0.85)
        self.start_button.place(relx=0.66, rely=0.85)

    def forward(self):
        """Linked to Information Sheet forward button.
        Places the consent page."""

        self.page1.place_forget()  # Function to present consent info
        self.page2.place_forget()
        self.show_consent_page()

    def back(self):
        """Linked to the Information Sheet back button.
        Removes the consent page and re-places the information sheet.
        Possibly has never been used by an actual participant."""

        self.consent_page.place_forget()  # Function to return to the information sheet
        self.information_page()

    def consent_check(self):
        """Linked to consent page start_button.
        Checks that ALL of the consent buttons are ticked.
        If they are not and the button is pressed, then the warning label is placed."""

        self.window.focus_set()  # Checks that all the consent boxes are ticked
        if self.read.get() == 1 and self.questions.get() == 1 and self.withdraw.get() == 1 and self.agree.get() == 1 and self.participate.get() == 1 and self.consent_var.get() == 1:
            # ParticipantResponses.append('Consent Given')
            self.demographics()
        elif self.read.get() != 1 or self.questions.get() != 1 or self.withdraw.get() != 1 or self.agree.get() != 1 or self.participate.get() != 1 or self.consent_var.get() != 1:
            self.warning_label.place(relx=0.5, rely=0.92, anchor=CENTER)

    def demographics(self):
        """Linked to the consent_check function.
        Brings up the demographic collection window for information collection."""

        self.consent_page.destroy()
        self.window.config(bg='white')
        self.window.attributes('-fullscreen', False)
        self.window.geometry(
            f'{int(WINDOW_WIDTH / 100) * 27}x{int(WINDOW_HEIGHT / 100) * 55}+{int(WINDOW_WIDTH / 2.5)}+{int(WINDOW_HEIGHT / 7.1)}')  # Places the window in the middle of the screen.
        self.window.title('Demographic Information')
        self.window.resizable(width=False, height=False)

        RESPONSES.append('Consent Given')

        # -------------------------------- Places the Demographics Window Variables --------------------------------
        self.entry_canvas.place(relx=0.5, rely=0.5, anchor=CENTER)
        self.demographics_frame.place(relx=0.5, rely=0.47, anchor=CENTER)
        self.demographics_label.place(relx=0.5, rely=0.16, anchor=CENTER)
        self.id_label.place(relx=.5, rely=.35, anchor=CENTER)
        self.age.place(relx=.5, rely=.50, width=int((WINDOW_WIDTH / 100) * 16.1),
                       height=int((WINDOW_HEIGHT / 100) * 3.81),
                       anchor=CENTER)
        self.email.place(relx=.5, rely=.65, width=int((WINDOW_WIDTH / 100) * 16.1),
                         height=int((WINDOW_HEIGHT / 100) * 3.81),
                         anchor=CENTER)
        self.gender.place(relx=.5, rely=.785, anchor=CENTER)
        self.submit_button.place(relx=0.5, rely=0.9, anchor=CENTER)

        self.age.insert(0, 'Enter Your Age')
        self.email.insert(0, 'Enter Your Email')
        self.age.bind("<1>", self.age_insert_clear)  # Bind clicking into the age entry box to clearing the box.
        self.email.bind('<1>', self.email_insert_clear)  # Bind clicking into the email entry box to clearing the box.

    def age_insert_clear(self, _):
        """Linked to the age entry box.
        If the contents of the age entry box appear in illegals, then delete it."""
        if self.age.get() in self.illegals:  # If age entry in illegals, removes entry.
            self.age.delete(0, END)

    def email_insert_clear(self, _):
        """Linked to the email entry box.
        If the contents of the email entry box appear in illegals, then delete it."""

        if self.email.get() in self.illegals:  # If email entry in illegals, removes entry.
            self.email.delete(0, END)

    def submit(self):  # Demographic entry checks
        """Linked to submit button.
        Checks all of the variables and if they're appropriate then starts the procedure with the VVIQ."""

        if self.age.get() in self.illegals:
            self.age.delete(0, END)
            self.age.insert(0, 'Age Required')
        elif int(self.age.get()) < 18:
            self.age.delete(0, END)
            self.age.insert(0, 'Invalid Age')
        elif self.email.get() in self.illegals:
            self.email.delete(0, END)
            self.email.insert(0, 'Email Required')
        elif '@' not in self.email.get():
            self.email.delete(0, END)
            self.email.insert(0, 'Valid Email Required')
        elif int(
                self.age.get()) >= 18 and '@' in self.email.get():  # If entries are appropriate, append responses to list
            RESPONSES.append(time.strftime('%X %Y%m%d'))  # Add the date and time to the master_responses list.
            RESPONSES.append(ID_CODE)
            RESPONSES.append(self.age.get())
            RESPONSES.append(self.selection.get())
            RESPONSES.append(self.email.get())

            self.demographics_frame.destroy()  # Destroys the canvases, Pretty sure I don't need this one.
            self.entry_canvas.destroy()

            # if THIS_RUN_TEST == 'VPT':
            #     Visual = VisualPatternTest(root, 2, False, COMICS, PRACTICE_COMICS, QUESTIONS)
            #     Visual.start()
            # elif THIS_RUN_TEST == 'Corsi':


            # elif THIS_RUN_TEST == 'Digit':
            #     Digit = DigitSpan(root, 3, False, COMICS, PRACTICE_COMIC, QUESTIONS)
            #     Digit.start()

            CorsiBlockTest(root, 2, False).setup()

class CorsiBlockTest:

    def __init__(self, window, level, dual_task):
        self.window = window
        if not DEV_MODE:
            self.window.attributes('-fullscreen', True)
        self.window.geometry(f'{WINDOW_WIDTH}x{WINDOW_HEIGHT}+{int(SCREEN_HEIGHT / 1)}+0')
        self.level = level
        self.dual_task = dual_task
        self.practice = True
        if EXPERIMENT:
            self.practice_comic = PRACTICE_COMIC
            self.comics = COMICS
            self.question_list = list(QUESTIONS)
            self.comic_counter = 0
            self.time_q = 'Approximately how much time passed in the comic?'
            self.instructions = [
                'In this part of the experiment you will see a number of randomly placed blocks on the screen.\n\nThe blocks will flash in a certain order and your task is to click the blocks in the order they flashed.\n\nThe number of blocks will slowly increase over time.\n\n\n\nClick start to begin the practice trials.',
                'Thank you for completing this questionnaire.\n\n\n\nThis is the final part of the experiment.\n\nIn this part you will complete the same type of block sequence trials, there will be a secondary task inbetween seeing and repeating the sequence.\n\n\n\nThe first trial is a practice trial.\n\nClick start to begin the practice trial.',
                'Thank you, the practice trial is now complete.\n\n\nDoes it make sense what you have to do?\n\n\nClick start to begin.']
        else:
            self.instructions = [
                'In this part of the experiment you will see a number of randomly placed blocks on the screen.\n\nThe blocks will flash in a certain order and your task is to click the blocks in the order they flashed.\n\nThe number of blocks will slowly increase over time.\n\n\n\nClick start to begin the practice trials.',
                'Thank you for completing this questionnaire.\n\n\n\nThis is the final part of the experiment.\n\nIn this part you will complete the same type of block sequence trials, but this time the gap between seeing and repeating the sequence will be longer.\n\n\n\nThe first trial is a practice trial.\n\nClick start to begin the practice trial.',
                'Thank you, the practice trial is now complete.\n\n\nDoes it make sense what you have to do?\n\n\nClick start to begin.']

        self.block_list = []
        self.response_list = []
        self.practice_list = []

        self.titration_trial_list = []
        self.intra_trial_outcomes = []
        self.master_trial_outcomes = []
        self.dual_task_responses = []

        self.dual_task_counter = 0

    def click_block(self, block):
        """Function that handles when a block is clicked during the Corsi Block presentation.
        Changes the colour of the block from red to black or vice versa.
        Adds the position number to the block if it's red.
        Also maintains the size of the block because it was increasing with every click for some reason."""
        if block['bg'] == 'black':
            self.response_list.append(block)
            block.config(text=f'{self.response_list.index(block) + 1}',
                         font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 17))), fg='White')
            block['width'] = int(WINDOW_WIDTH * 0.05)
            block['height'] = int(WINDOW_WIDTH * 0.05)
            block['bg'] = 'red'
            root.update()
        elif block['bg'] == 'red':
            self.response_list.remove(block)
            block.config(text='')
            block['bg'] = 'black'
            block['width'] = int(WINDOW_WIDTH * 0.05)
            block['height'] = int(WINDOW_WIDTH * 0.05)
            root.update()
        root.focus_set()

    def setup(self):
        """Function handles the beginning of any Corsi Block trial with stipulations for
        certain conditions under which instructions need to be shown."""
        self.start_button = Button(self.window, text='Start', command=self.pre_trial_fixation_cross)
        self.instructions_label = Label(self.window, bg='White', highlightthickness=0, bd=0)
        # Titration Pre Instructions
        if not self.dual_task and len(self.titration_trial_list) == 0:
            self.instructions_label.config(text=self.instructions[0], wraplength=800,
                                           font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15))),
                                           justify='center')
            self.instructions_label.place(relx=.5, rely=.5, anchor=CENTER)
            self.start_button.place(relx=0.5, rely=0.85, anchor=CENTER)
        # Titration Post Instructions
        elif len(self.titration_trial_list) == 2:
            self.instructions_label.config(text=self.instructions[2], wraplength=800,
                                           font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15)),),
                                           justify='center')
            self.instructions_label.place(relx=.5, rely=.5, anchor=CENTER)
            self.start_button.place(relx=0.5, rely=0.85, anchor=CENTER)
        # Dual Task Pre Instructions
        elif self.dual_task and len(self.practice_list) == 0:
            self.instructions_label.config(text=self.instructions[1], wraplength=800,
                                           font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15))),
                                           justify='center')
            self.instructions_label.place(relx=.5, rely=.5, anchor=CENTER)
            self.start_button.place(relx=0.5, rely=0.85, anchor=CENTER)
        # Dual Task Post Instructions
        elif self.practice and len(self.practice_list) == 1: # Dual Task Post Instructions
            try:
                self.time_question_label.destroy()
                self.time_entry.destroy()
            except:
                pass
            self.instructions_label.config(text=self.instructions[2], wraplength=800,
                                           font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15))),
                                           justify='center')
            self.instructions_label.place(relx=.5, rely=.5, anchor=CENTER)
            self.start_button.place(relx=0.5, rely=0.85, anchor=CENTER)
        else:
            self.start_button.place(relx=.5, rely=.5, anchor=CENTER)

    def pre_trial_fixation_cross(self):
        """The initial part of any Corsi trial, the interstimulus cross"""
        self.start_button.destroy()
        self.instructions_label.destroy()
        self.cross = Label(self.window, text='+', bg='White',
                           font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 65))))
        self.cross.place(relx=.5, rely=.5, anchor=CENTER)
        self.window.after(two_seconds, self.presentation)

    def presentation(self):
        """Linked to interstimulus_cross().
        Creates blocks based on level, and sets their function to click_block() above.
        Creates random positions for the blocks and places them on the screen.
        Sets flash_list as the blocks that will flash.
        Sets the block position based on the positions created using indexes.
        """
        self.cross.destroy()
        self.window.update()
        self.window.config(cursor='none')
        self.blocks_canvas = Canvas(self.window, width=WINDOW_WIDTH + 10, height=WINDOW_HEIGHT + 10, highlightthickness=0, bg='White', bd=0)
        self.blocks_canvas.pack()
        for x in range(0, self.level):
            self.blocks = Button(self.blocks_canvas, width=int(WINDOW_WIDTH * 0.05),
                                 height=int(WINDOW_WIDTH * 0.05),
                                 bg='black', highlightthickness=4, focuscolor='', fg='black')
            self.blocks.config(command=lambda block_ref=self.blocks: self.click_block(block_ref))
            self.block_list.append(self.blocks)
        self.response_list = []
        self.cross.destroy()
        self.flash_list = [x for x in self.block_list if self.block_list.index(x) < self.level]  # If the index of the block is less than self.level, then put that block in a list for flashing.
        self.x = random.sample(range(9, 91, 10), 9)
        self.y = random.sample(range(9, 91, 10), 9)  # Creates random positions for the blocks on screen the blocks in random places on the screen.
        self.x_i = 0
        self.y_i = 0
        for block in self.block_list:
            block.place(relx=self.x[self.x_i] / 100, rely=self.y[self.y_i] / 100,
                        anchor=CENTER)  # Sets the block position based on the positions created using indexes.
            self.x_i += 1
            self.y_i += 1
        root.after(two_seconds, self.flash_routine)

    def flash_routine(self):
        """Linked to presentation().
        For every block in flash list, turn it red, wait a second, turn it black, wait 0.2 seconds
        Repeat this process until it has happened to every block."""
        self.window.config(cursor='none')
        for block in self.flash_list:
            if DEV_MODE or FAST_COMPONENT_DEMO:
                block.config(bg='red')
                root.update()
                time.sleep(0.2)  # 1
                block.config(bg='black')
                root.update()
                time.sleep(0.2)  # 0.2
            else:
                block.config(bg='red')
                root.update()
                time.sleep(1)  # 1
                block.config(bg='black')
                root.update()
                time.sleep(0.2)  # 0.2
        root.after(two_seconds, self.trial_selection)

    def trial_selection(self):
        """Linked to flash_routine().
        Set up a canvas for that covers the blocks post flash routine.
        If dual_task is not turned on, go straight to titration_trial().
        If dual task is not turned on, go pick_trial()."""
        self.window.config(cursor='none')
        self.blocks_canvas.pack_forget()
        self.window.update()

        if not self.dual_task:
            if len(self.titration_trial_list) >= 2:
                self.practice = False
            self.titration_trial()

        elif self.dual_task:
            if EXPERIMENT:
                if len(self.practice_list) == 1:
                    self.practice = False
                self.dual_task_trial()
            elif CONTROL:
                self.titration_trial()

    def titration_trial(self):
        """Linked from initial_trial_choice().
        Add 1 to titration trial.
        Show the label "Repeat the Sequence".
        Go to response()."""
        self.titration_instruction_canvas = Canvas(self.window, width=WINDOW_WIDTH + 10, height=WINDOW_HEIGHT + 10,
                                                   highlightthickness=0, bg='White')
        self.titration_instruction_canvas.pack()
        self.window.config(cursor='none')
        if not self.dual_task:
            self.label = Label(self.titration_instruction_canvas, text='Repeat the Sequence',
                               font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15))), bg='White')  # Basic Trial
            self.label.place(relx=.5, rely=.5, anchor=CENTER)
            self.titration_trial_list.append(1)
            self.window.after(two_seconds, self.response_screen)
        elif self.dual_task:
            if self.practice:
                self.practice_list.append(1)
            self.label = Label(self.titration_instruction_canvas, text='+',
                               font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15))), bg='White')  # Basic Trial
            self.label.place(relx=.5, rely=.5, anchor=CENTER)
            self.window.after(seven_seconds, self.response_screen)
            self.dual_task_counter += 1

    def dual_task_trial(self):
        if self.practice:
            self.practice_comic.place(relx=.5, rely=.5, anchor=CENTER)
            self.practice_list.append(1)
        else:
            self.comics[self.comic_counter].place(relx=.5, rely=.5, anchor=CENTER)
            self.window.update()
            RESPONSES.append(COMIC_NAMES[self.comic_counter])
        self.window.after(seven_seconds, self.response_screen)

    def response_screen(self):
        """Linked from titration_trial(), comic_trial(), control_trial(), empty_trial().
        Destroys the canvas that the interference was presented on bringing back the Corsi Blocks for response.
        Blocks during this phase are activated and deactivated using click_blocks()
        Unbinds mouse button if used, binds space bar to entering response."""
        if not self.dual_task: # if titration
            self.titration_instruction_canvas.destroy()
        elif self.dual_task and self.practice and EXPERIMENT:
            self.practice_comic.destroy()
        elif self.dual_task and not self.practice and EXPERIMENT:
            self.comics[self.comic_counter].destroy()
            self.comic_counter += 1
        elif self.dual_task and CONTROL:
            self.titration_instruction_canvas.destroy()
        self.blocks_canvas.pack()  # Put blocks back on screen in clickable format for response.
        self.window.config(cursor='arrow')
        self.window.update()
        root.bind_all('<space>', lambda x: self.response_check())

    def response_check(self):
        """Linked to response().
        Registers the response input by the participant.
        Unbinds the space bar from entering responses.
        If dual_task is FALSE (during titration phase):
            If the response_list of buttons pressed is equal to the flash list, that is, they are exactly the same, in order, then add 1 to intra_trial_outcomes.
            Otherwise add 0 to intra_trial_outcomes.

            If there have been 2 trials, and the sum of the intra_trial_outcomes list is 2 then go to level_up().
            If there have been MORE THAN 2 trials, and the sum of the intra_trial_outcomes list is 3 then go to level_up().
            If there are two zeroes in the intra_trial_outcomes list then:
                append the number of titration_trials to master_responses
                append the current level -1 which will set the number of blocks for the dual task phase.
                destroy all relevant canvases and blocks,
                Start the VLFI sequence.

        If dual task is TRUE:
            Create the trial_target sequence.
            Go through the trial response list and:
                Add the elements of the list which match the elements of trial_target in both position and data.
                Make trial_response the same length as trial_target if it isn't.

            Determine the accuracy of the response and append that to master_responses.
            Initialise variables for, and call the question phase."""
        root.unbind_all('<space>')
        self.blocks_canvas.destroy()
        # if not in dual task mode,
        if not self.dual_task:
            if self.practice and len(self.titration_trial_list) < 2:
                self.next_trial()
            elif self.practice and len(self.titration_trial_list) == 2:
                self.level_up()
            # Check Response
            else:
                if self.response_list == self.flash_list:
                    self.intra_trial_outcomes.append(1)
                else:
                    self.intra_trial_outcomes.append(0)

                if len(self.intra_trial_outcomes) == 2 and sum(self.intra_trial_outcomes)== 2:
                    self.level_up()
                elif len(self.intra_trial_outcomes) > 2 and sum(self.intra_trial_outcomes) == 3:
                    self.level_up()
                elif self.intra_trial_outcomes.count(0) == 2:
                    self.master_trial_outcomes.append(self.intra_trial_outcomes[:])
                    RESPONSES.extend([f'Corsi Trials: {len(self.titration_trial_list) - 2}', int(self.level - 1)])
                    self.block_list.clear()
                    vlfi = VisualLanguageFluencyIndex(root)
                    vlfi.start()
                else:
                    self.next_trial()

        elif self.dual_task:
            self.trial_target = [str(n) for n in range(2, self.level + 1)]
            self.trial_target.insert(0, 'n')

            self.trial_response = [(str(n)[-1:]) for n in self.response_list]

            while len(self.trial_response) < len(self.trial_target):
                self.trial_response.append('-')
            while len(self.trial_response) > len(self.trial_target):
                self.trial_response.pop()

            self.trial_outcome = [response for response, target in zip(self.trial_response, self.trial_target) if
                                  response == target]
            self.trial_outcome_percent = float(len(self.trial_outcome) / len(self.trial_target))

            RESPONSES.extend([f'Target: {str(self.trial_target)} Response: {str(self.trial_response)}', self.trial_outcome_percent])

            if EXPERIMENT:
                self.time_question_display()
            elif CONTROL:
                if self.dual_task_counter == 6:
                    self.finish_procedure()
                else:
                    self.next_trial()

    def time_question_display(self):
        # self.time_question_canvas = Canvas(root, width=WINDOW_WIDTH+3, height=WINDOW_HEIGHT+3, bg='white')
        self.time_question_label = Label(root, text=self.time_q, background='white', font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 20))))
        self.time_entry = Entry(root, justify='center', relief='sunken', width=20, font=('default', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 20))))
        self.time_question_label.place(relx=.5, rely=.35, anchor=CENTER)
        self.time_entry.place(relx=.5, rely=.5, anchor=CENTER)
        self.time_entry.focus_set()
        # self.time_question_canvas.pack()
        self.time_entry.bind('<Return>', self.time_question_handler)

    def time_question_handler(self, _):
        self.time_entry.unbind('<Return>')
        if not self.practice:
            RESPONSES.append(self.time_entry.get())
            self.dual_task_counter += 1
        self.time_question_label.destroy()
        self.time_entry.destroy()
        if self.dual_task_counter == 6:
            QuestionPhase(root).instructions()
        else:
            self.next_trial()

    def level_up(self):
        """Linked from response_check() if the user completes a set of titration trials successfully.
        Increases the level variable, resets intra_trial_outcomes list to start again and appends the intra_trial_outcomes to master_trial_outcomes.
        Goes to next_trial()."""
        if not self.practice:
            self.master_trial_outcomes.append(self.intra_trial_outcomes[:])
        self.level += 1
        self.intra_trial_outcomes.clear()
        self.next_trial()

    def next_trial(self):
        self.block_list.clear()
        self.setup()

    def finish_procedure(self):
        try:
            self.instructions_label.destroy()
        except:
            pass
        self.end_label = Label(root,
                               text='Thank you! Your participation is really appreciated!\n\nThis experiment is '
                                    'about mental model creation, visuospatial working memory and comics reading.\nEach type of trial that '
                                    'you completed is intended to demonstrate a different cognitive element of '
                                    'creating mental models from reading comics.\n\nIf you have any questions, please feel free to ask them.\nYou '
                                    'should now have received an email containing follow up information.\n\n Thanks '
                                    'again!',
                               wraplength=800, font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15))),
                               justify='center', bg='White')
        self.end_button = Button(self.window, text='Complete', command=quit_protocol)
        save_data(DATA_WORKBOOK, RESPONSES)
        self.end_label.place(relx=.5, rely=.5, anchor=CENTER)
        self.end_button.place(relx=.5, rely=.8, anchor=CENTER)
        self.window.update()
        time.sleep(1)

class VisualLanguageFluencyIndex:
    """This class handles everything to do with the VLFI throughout this experiment."""

    def __init__(self, window):
        # Basic Window Setup
        self.window = window
        if not DEV_MODE:
            self.window.attributes('-fullscreen', True)
        self.canvas = Canvas(self.window, width=SCREEN_WIDTH + 10, height=SCREEN_HEIGHT + 10, highlightthickness=0,
                             bg='White')
        self.canvas.pack()

        # Counters
        self.q_set_i = 0  # Used to handle which question used.
        self.r_set_i = 0  # Used to hand the response set used.

        # Lists
        self.q_set = [['On average, how often per week do you read text-only books for enjoyment?',  # 0
                       'On average, how often per week do you watch films?',  # 1
                       'On average, how often per week do you watch cartoons/anime?',  # 2
                       'On average, how often per week do you read comic books?',  # 3
                       'On average, how often per week do you read comic strips?',  # 4
                       'On average, how often per week do you read graphic novels?',  # 5
                       'On average, how often per week do you read Japanese comics (manga)?',  # 6
                       'On average, how often per week do you draw comics?'],
                      ['On average, how often per week did you read text-only books for enjoyment while growing up?',
                       # 8
                       'On average, how often per week did you watch films while growing up?',  # 9
                       'On average, how often per week did you watch cartoons/anime while growing up?',  # 10
                       'On average, how often per week did you read comic books while growing up?',  # 11
                       'On average, how often per week did you read comic strips while growing up?',  # 12
                       'On average, how often per week did you read graphic novels while growing up?',  # 13
                       'On average, how often per week did you read Japanese comics (manga) while growing up?',  # 14
                       'On average, how often per week did you draw comics while growing up?'],
                      ['How would you currently rate your expertise with reading comics (of any sort)?',  # 16
                       'How would you currently rate your drawing ability?',
                       'How would you rate your expertise with reading comics (of any sort) while growing up?',  # 18
                       'How would you rate your drawing ability while growing up?']]  # List of questions as strings.
        self.r_set = [[('1', 1),
                       ('2', 2),
                       ('3', 3),
                       ('4', 4),
                       ('5', 5),
                       ('6', 6),
                       ('7', 7)], [('1', 1),
                                   ('2', 2),
                                   ('3', 3),
                                   ('4', 4),
                                   ('5', 5)]]  # List of responses.
        self.e_set = ['Approximately how old were you when you first read a comic?',
                      'Approximately how old were you when you first started drawing comics? (0 = N/A or Never)']  # Drop down questions.
        self.responses = []  # List to handle responses.

        self.instruction_label = None
        self.start_button = None
        self.q_l = None
        self.i = None
        self.question_rdb_canvas = None
        self.label = None
        self.rdb = None
        self.submit_button = None
        self.age_options = None
        self.values = None
        self.question_canvas = None
        self.ages = None
        self.indices = None
        self.rc = None
        self.a = None
        self.b = None
        self.c = None
        self.d = None
        self.e = None
        self.f = None
        self.g = None
        self.h = None
        self.a = None
        self.vlfi_score = None

    def start(self):
        """Begins the VLFI procedure.
         Presents an instruction and a button to start the procedure."""
        self.instruction_label = Label(self.canvas,
                                       text='Thank you, this part of the experiment is now complete.\n\nPlease complete the following short questionnaire on your familiarity with comics.',
                                       wraplength=800, font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15)),),
                                       justify='center', bg='White')
        self.instruction_label.place(relx=.5, rely=.5, anchor=CENTER)
        self.start_button = Button(self.canvas, text='Start', command=self.questions)
        self.start_button.place(relx=.5, rely=.8, anchor=CENTER)

    def questions(self):
        """Linked to start_button in start().
        This function presents quesitons to the user and is used recursively to present all of the question sets.
        Sets the values of each of the questions to an IntVar()
        Then loads the questions from a list onto a canvas and applies a label to them.
        For each of the questions set radiobuttons.
        Place the submit button."""
        self.start_button.destroy()
        self.instruction_label.destroy()
        self.q_l = []
        self.question_canvas = Canvas(self.canvas, highlightthickness=0, borderwidth=0, bg='White')
        self.values = [IntVar() for _ in
                       range(len(self.q_set[self.q_set_i]))]  # Sets the values of each of the questions to an IntVar()
        self.i = 0
        for n in self.q_set[self.q_set_i]:
            self.question_rdb_canvas = Canvas(self.question_canvas, bg='White')
            self.label = Label(self.question_canvas, text=n,
                               font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))),
                               pady=int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14)), bg='White')
            self.label.pack()
            self.q_l.append(self.label)
            for t, v in self.r_set[self.r_set_i]:
                self.rdb = Radiobutton(self.question_rdb_canvas, text=t, val=v, variable=self.values[self.i],
                                       font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))),
                                       padx=5, bg='White').pack(side=LEFT)
            self.question_rdb_canvas.pack()
            self.i += 1
        self.submit_button = Button(self.question_canvas, text='Submit', command=self.submit).pack(pady=20)
        self.question_canvas.place(relx=.5, rely=.5, anchor=CENTER)

    def entry_questions(self):
        """This function handles the entry_questions. (Which are no longer entry questions but dropdown menus.
        It does this in a similar way to the quesitons() function."""
        self.i = 0
        self.age_options = [x for x in range(0, 66)]
        self.values = [IntVar() for _ in range(len(self.e_set))]
        self.question_canvas = Canvas(self.canvas, highlightthickness=0, borderwidth=0, bg='White')
        for n in self.e_set:
            self.question_rdb_canvas = Canvas(self.question_canvas, bg='White')
            self.ages = OptionMenu(self.question_canvas, self.values[self.i], *self.age_options)
            self.label = Label(self.question_canvas, text=n,
                               font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 14))), bg='White')
            self.label.pack(pady=15)
            self.ages.pack(pady=15)
            self.i += 1
        self.question_canvas.place(relx=.5, rely=.5, anchor=CENTER)
        self.submit_button = Button(self.question_canvas, text='Submit', command=self.submit).pack(pady=20)

    def submit(self):
        """Linked to questions() and entry_questions()
        This function gets all the responses from the radiobuttons presented on the previous page.
        If there is a 0 in the responses list then highlight that question.
        Otherwise, remove the canvas, add the responses to the necessary list
        then decide on where to go from there.
        If all questions have been answered go to calculate()."""

        self.rc = [x.get() for x in self.values]
        if len(self.responses) < 20 and 0 in self.rc:
            self.indices = [i for i, x in enumerate(self.rc) if x == 0]
            for n in self.indices:
                self.q_l[n].configure(fg='red')
                self.canvas.update()
        elif len(self.responses) != 20:
            self.question_canvas.destroy()
            self.responses.extend(self.rc)
            self.q_set_i += 1
            if self.q_set_i > 2:
                self.entry_questions()
            elif self.q_set_i < 2:
                self.questions()
            elif self.q_set_i == 2:
                self.r_set_i += 1
                self.questions()
        else:
            self.responses.extend(self.rc)
            self.calculate()

    def calculate(self):
        """Linked to submit().
        This function handles Cohn's ridiculous equation for calculating the VLFI score.
        It then appends this item to the master_responses list
        Finally, it calls an instance of the dual task class which sets the dual_task argument to true."""
        self.question_canvas.destroy()
        self.a = sum(self.responses[3:7]) / 4
        self.b = sum(self.responses[10:14]) / 4
        self.c = self.a * self.responses[16]
        self.d = self.b * self.responses[18]
        self.e = (self.c + self.d) / 2
        self.f = (self.responses[7] + self.responses[15]) / 2
        self.g = (self.responses[17] + self.responses[19]) / 2
        self.h = self.f * self.g / 2
        self.vlfi_score = self.e + self.h

        RESPONSES.append(self.vlfi_score)
        self.canvas.destroy()
        self.question_canvas.destroy()
        self.question_rdb_canvas.destroy()

        CorsiBlockTest(root, RESPONSES[8], True).setup()

class QuestionPhase:

    def __init__(self, window):
        self.window = window

        self.window.geometry(f'{WINDOW_WIDTH}x{WINDOW_HEIGHT}+{int(SCREEN_HEIGHT / 1)}+0')
        self.content_canvas = Canvas(root, bg='white', bd=0, highlightthickness=0)
        self.panels_canvas = Canvas(root, bg='white', bd=0, highlightthickness=0)
        self.content_prompt = Label(self.content_canvas, text='What happened in the comic?',
                                    font=('default', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 20))), bg='white')
        self.panels_prompt = Label(self.panels_canvas, text='No. of panels?',
                                   font=('default', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 20))), bg='white')
        self.content_prompt.pack()
        self.panels_prompt.pack()
        self.description_entries_list = []
        self.panel_entries_list = []
        for _ in range(0, 6):
            self.descriptions_var = StringVar()
            self.descriptions_entry_box = Entry(self.content_canvas, relief='sunken', borderwidth=2,
                                                font=("Arial", int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 24))),
                                                width=int(WINDOW_WIDTH * 0.015), textvariable=self.descriptions_var)
            self.description_entries_list.append(self.descriptions_entry_box)
            self.descriptions_entry_box.pack(pady=10)
        for _ in range(0, 6):
            self.panels_var = StringVar()
            self.panel_entry_box = Entry(self.panels_canvas, justify='center', relief='sunken', borderwidth=2,
                                         font=("Arial", int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 24))),
                                         width=int(WINDOW_WIDTH * 0.003), textvariable=self.panels_var)
            self.panel_entries_list.append(self.panel_entry_box)
            self.panel_entry_box.pack(pady=10)
        self.question_counter = 5
        self.free_recall_instructions = """On the following page, you'll be asked to describe what happened in the comics you saw.\n\nYou should exclude the award comic that you saw in the practice trial.\n\n\nYou can do this in any order you like.\n\n Next to each of the description boxes, please indicate how many panels the comic strip had.\n\nDon't worry about spelling or grammar here, try not to think about your answers too much,\nand write what comes naturally. If you can't remember, take a guess.\n\nOnce you submit your answers, there will be 6 more questions about the comics and then the experiment will finish."""

    def instructions(self):
        self.instructions_label = Label(root, text=self.free_recall_instructions,
                                        font=('default', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 20))), bg='white')
        self.instructions_label.place(relx=.5, rely=.5, anchor=CENTER)
        self.start_button = Button(root, text='Next', command=self.free_recall)
        self.start_button.place(relx=.5, rely=.75, anchor=CENTER)

    def free_recall(self):
        self.start_button.place_forget()
        self.instructions_label.place_forget()
        self.content_canvas.place(relx=.45, rely=.5, anchor=CENTER)
        self.panels_canvas.place(relx=.7, rely=.5, anchor=CENTER)
        self.window.update()
        self.window.after(ten_seconds, self.place_button_1)

    def place_button_1(self):
        self.start_button.config(command=self.get_responses, text='Next')
        self.start_button.place(relx=.5, rely=.85, anchor=CENTER)
        self.window.update()

    def get_responses(self):
        self.content_prompt.destroy()
        self.panels_prompt.destroy()
        self.descriptions = [entry.get() for entry in self.description_entries_list]
        self.panel_responses = [entry.get() for entry in self.panel_entries_list]
        self.tuple_list = list(zip(self.descriptions, self.panel_responses))
        self.free_recall_responses = [e for l in self.tuple_list for e in l]
        self.start_button.place_forget()
        self.content_canvas.destroy()
        self.panels_canvas.destroy()
        self.specific_questions()

    def specific_questions(self):
        self.start_button.config(command=self.next_q, text='Next')
        self.start_button.place(relx=.5, rely=.85, anchor=CENTER)
        self.instructions_label.config(text=QUESTIONS[self.question_counter])
        self.instructions_label.place(relx=.5, rely=.5, anchor=CENTER)
        self.question_answer = StringVar()
        self.question_answer_entry_box = Entry(root, justify='center', relief='sunken', borderwidth=2,
                                               font=('default', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 20))),
                                               textvariable=self.question_answer)
        self.question_answer_entry_box.place(relx=.5, rely=.65, anchor=CENTER)
        self.window.update()

    def next_q(self):
        RESPONSES.append(f'{QUESTIONS[self.question_counter]} : {self.question_answer.get()}')
        self.question_answer_entry_box.destroy()
        self.start_button.place_forget()
        if self.question_counter == 0:
            self.finish_procedure()
        else:
            self.question_counter -= 1
            self.specific_questions()

    def finish_procedure(self):
        self.instructions_label.destroy()
        self.end_label = Label(root,
                               text='Thank you! Your participation is really appreciated!\n\nThis experiment is '
                                    'aboout mental model creation, visuospatial working memory and comics reading.\nEach type of trial that '
                                    'you completed is intended to demonstrate a different cognitive element of '
                                    'creating mental models from reading comics.\n\nIf you have any questions, please feel free to ask them.\nYou '
                                    'should now have received an email containing follow up information.\n\n Thanks '
                                    'again!',
                               wraplength=800, font=('Arial', int(WINDOW_HEIGHT / font(SCREEN_HEIGHT, 15))),
                               justify='center', bg='White')
        self.end_button = Button(self.window, text='Complete', command=quit_protocol)
        if not DEV_MODE and not FULL_COMPONENT_DEMO and not FAST_COMPONENT_DEMO:
            try:
                email_protocol()
            except:
                pass
        save_data(DATA_WORKBOOK, RESPONSES)
        self.end_label.place(relx=.5, rely=.5, anchor=CENTER)
        self.end_button.place(relx=.5, rely=.8, anchor=CENTER)
        self.window.update()
        time.sleep(1)

if DEV_MODE or FAST_COMPONENT_DEMO or FULL_COMPONENT_DEMO:
    component = input('\nChoose component to run:\n   0 - Full\n   1 - Information Sheet\n   2 - Corsi Block Titration\n   3 - VLFI\n   4 - Corsi Block Dual Task\n   5 - Question Phase\n>>> ')

    components_dict = {
        '0': lambda: InformationSheet(root).information_page(),
        '1': lambda: InformationSheet(root).information_page(),
        '2': lambda: CorsiBlockTest(root, 2, False).setup(),
        '3': lambda: VisualLanguageFluencyIndex(root).start(),
        '4': lambda: CorsiBlockTest(root, 5, True).setup(),
        '5': lambda: QuestionPhase(root).instructions()
    }
    if component in components_dict:
        components_dict[component]()
        root.mainloop()

else:
    InformationSheet(root).information_page()
    root.mainloop()